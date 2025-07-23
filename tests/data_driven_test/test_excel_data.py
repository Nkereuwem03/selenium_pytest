from time import time
from utils.logger import logger
import pytest
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
)
from utils.wait_helper import wait_for_element_presence
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from utils.paths import get_absolute_path
from utils.excel_reader import get_row_count, update_cell, load_sheet
from openpyxl.styles import PatternFill
from selenium.webdriver.remote.webdriver import WebDriver


EXCEL_PATH = get_absolute_path("data", "excel_data.xlsx")
GREEN_FILL = PatternFill(start_color="60b212", end_color="60b212", fill_type="solid")
RED_FILL = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")


def find_element(driver: WebDriver, locator: tuple, data):
    element = wait_for_element_presence(driver, locator)
    element.clear()
    element.send_keys(data)


def test_fixed_deposit_calculator(setup_teardown):
    driver = setup_teardown
    max_retries = 3
    for attempt in range(max_retries):
        try:
            driver.set_page_load_timeout(60)
            driver.get("https://fd-calculator.in/result")
            break 
        except TimeoutException:
            if attempt == max_retries - 1:
                pytest.fail("Failed to load page after multiple attempts")
            logger.warning(f"Page load timeout, attempt {attempt + 1}, retrying...")
            time.sleep(5)  
    file = EXCEL_PATH

    sheet = load_sheet(file, "Sheet3")

    rows = get_row_count(file, "Sheet3")

    try:

        # for row in range(1, rows + 1):
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            # principal = read_data(file, sheet, row, 1)
            # rate = read_data(file, sheet, row, 2)
            # period = read_data(file, sheet, row, 3)
            # period_unit = read_data(file, sheet, row, 4)
            # frequency = read_data(file, sheet, row, 5)
            # expected_value = read_data(file, sheet, row, 6)

            amount = row[0]
            period_value = row[1]
            period_unit = row[2]
            interest = row[3]
            frequency = row[4]
            expected_value = row[5]
            
            print(amount, period_value, period_unit, interest, frequency, expected_value)

            # if(any(value is None or str(value).strip() == "" for value in [principal, rate, period, period_unit, frequency, expected_value])):
            #     logger.warning(f"Row {row}: Skipping due to empty/None values")
            #     write_data(file, sheet, row, 8, "skipped")
            #     continue

            logger.info(f"Processing row {row}")

            find_element(driver, (By.CSS_SELECTOR, "#amountInputField"), str(amount))
            find_element(driver, (By.CSS_SELECTOR, "#periodInputField"), str(period_value))

            try:
                Select(
                    driver.find_element(By.ID, "amountSelectField")
                ).select_by_visible_text(str(period_unit).strip())
            except (NoSuchElementException, TimeoutException, ValueError) as e:
                logger.error(f"Test failed due to: {type(e).__name__}: {e}")
                update_cell(file, "Sheet3", idx, 8, value="fail", fill=RED_FILL)
                continue

            find_element(driver, (By.CSS_SELECTOR, "#interestInputField"), str(interest))

            try:
                Select(
                    driver.find_element(By.ID, "frequencySelectField")
                ).select_by_visible_text(str(frequency).strip())
            except (NoSuchElementException, TimeoutException, ValueError) as e:
                logger.error(f"Test failed due to: {type(e).__name__}: {e}")
                update_cell(file, "Sheet3", idx, 8, value="fail", fill=RED_FILL)
                continue

            try:
                calc_btn = driver.find_element(By.ID, "calculateButton")
                driver.execute_script("arguments[0].click();", calc_btn)
            except (NoSuchElementException, TimeoutException, ValueError) as e:
                logger.error(f"Test failed due to: {type(e).__name__}: {e}")
                continue

            try:
                result_elem = wait_for_element_presence(
                    driver, (By.CSS_SELECTOR, "#futureValue")
                )
                actual_value = float(result_elem.text.replace("Lakh", "").strip())
                expected_value = float(expected_value)

                if round(actual_value, 1) == round(expected_value, 1):
                    update_cell(file, "Sheet3", idx, 8, value="pass", fill=GREEN_FILL)
                else:
                    update_cell(file, "Sheet3", idx, 8, value="fail", fill=RED_FILL)

            except (NoSuchElementException, TimeoutException, ValueError) as e:
                logger.error(f"[Row {row}] Error during calculation: {e}")
                pytest.fail(f"Test failed due to: {type(e).__name__}: {e}")
                update_cell(file, "Sheet3", idx, 8, value="fail", fill=RED_FILL)

    except (NoSuchElementException, TimeoutException, ValueError) as e:
        logger.error(f"Test failed due to: {type(e).__name__}: {e}")
        pytest.fail(f"Test failed due to: {type(e).__name__}: {e}")
