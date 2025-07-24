"""Utility functions for reading from and writing to Excel files."""

import openpyxl
from openpyxl.styles import PatternFill


def load_sheet(file, sheet_name):
    """
    Loads a specified sheet from an Excel file.

    Args:
        file (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to load.

    Returns:
        openpyxl.worksheet.worksheet.Worksheet: The loaded sheet.
    """
    workbook = openpyxl.load_workbook(file)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")
    return workbook[sheet_name]


def get_row_count(file, sheet_name):
    """
    Gets the total number of rows in a specified Excel sheet.

    Args:
        file (str): The path to the Excel file.
        sheet_name (str): The name of the sheet within the Excel file.

    Returns:
        int: The number of rows in the sheet.
    """
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def read_data(file, sheet_name, row_num, col_num):
    """
    Reads data from a specific cell in an Excel sheet.

    Args:
        file (str): The path to the Excel file.
        sheet_name (str): The name of the sheet within the Excel file.
        row_num (int): The row number of the cell (1-indexed).
        col_num (int): The column number of the cell (1-indexed).

    Returns:
        any: The value from the specified cell.
    """
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    data = sheet.cell(row_num, col_num).value
    return data


def update_cell(file, sheet_name, row, col, value=None, fill=None):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    if value is not None:
        sheet.cell(row=row, column=col).value = value
    if fill is not None:
        sheet.cell(row=row, column=col).fill = fill
    workbook.save(file)
